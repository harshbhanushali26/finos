"""
FinOS — export routes (api/routes/export.py)

GET /export/csv    — download CSV
GET /export/json   — download JSON
GET /export/excel  — download .xlsx
GET /export/pdf    — download PDF

All stubs for now — export/exporters.py is built in Phase 5.
"""

import csv
import io
import json
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from api.deps import get_current_user, get_db
from core.models import Transaction, User

router = APIRouter(prefix="/export", tags=["export"])



def get_transactions(db: Session, user: User, month: Optional[str], year: Optional[int]):
    stmt = select(Transaction).where(Transaction.user_id == user.id)
    if month:
        # month format: "2026-06"
        try:
            y, m = month.split('-')
            from datetime import date as dt_date
            start = dt_date(int(y), int(m), 1)
            import calendar
            last_day = calendar.monthrange(int(y), int(m))[1]
            end = dt_date(int(y), int(m), last_day)
            stmt = stmt.where(Transaction.date >= start, Transaction.date <= end)
        except Exception:
            pass
    elif year:
        from datetime import date as dt_date
        stmt = stmt.where(
            Transaction.date >= dt_date(year, 1, 1),
            Transaction.date <= dt_date(year, 12, 31)
        )
    stmt = stmt.order_by(Transaction.date.desc())
    return db.exec(stmt).all()


@router.get("/csv")
def export_csv(
    month: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    txns = get_transactions(db, current_user, month, year)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Date", "Type", "Category", "Amount", "Note"])
    from datetime import datetime
    for t in txns:
        # writer.writerow([t.id, str(t.date), t.type, t.category, t.amount, t.note or ""])
        d = datetime.strptime(str(t.date), "%Y-%m-%d")
        writer.writerow([t.id, d.strftime("%d-%b-%Y"), t.type, t.category, t.amount, t.note or ""])
    output.seek(0)

    filename = f"finos-{month or year or 'all'}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/json")
def export_json(
    month: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    txns = get_transactions(db, current_user, month, year)
    data = [{"id": t.id, "date": str(t.date), "type": t.type, "category": t.category, "amount": t.amount, "note": t.note} for t in txns]
    content = json.dumps({"user": current_user.username, "count": len(data), "transactions": data}, indent=2)
    filename = f"finos-{month or year or 'all'}.json"
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel")
def export_excel(
    month: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    txns = get_transactions(db, current_user, month, year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # header style
    hdr_font = Font(bold=True, color="1E1B4B", size=11)
    hdr_fill = PatternFill("solid", fgColor="6366F1")
    hdr_align = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Date", "Type", "Category", "Amount (₹)", "Note"]
    col_widths = [8, 16, 12, 18, 16, 30]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # row styles
    inc_fill = PatternFill("solid", fgColor="D1FAE5")
    exp_fill = PatternFill("solid", fgColor="FEE2E2")

    for row_i, t in enumerate(txns, 2):
        values = [t.id, str(t.date), t.type, t.category, float(t.amount), t.note or ""]
        fill = inc_fill if t.type == "income" else exp_fill
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            if col_i == 5:  # amount column
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"finos-{month or year or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pdf")
def export_pdf():
    return {"detail": "PDF export coming soon"}

















# _NOT_IMPLEMENTED = JSONResponse(
#     status_code=501,
#     content={"detail": "Export not yet implemented — coming in Phase 5"},
# )


# @router.get("/csv")
# def export_csv(current_user: User = Depends(get_current_user)):
#     return _NOT_IMPLEMENTED


# @router.get("/json")
# def export_json(current_user: User = Depends(get_current_user)):
#     return _NOT_IMPLEMENTED


# @router.get("/excel")
# def export_excel(current_user: User = Depends(get_current_user)):
#     return _NOT_IMPLEMENTED


# @router.get("/pdf")
# def export_pdf(current_user: User = Depends(get_current_user)):
#     return _NOT_IMPLEMENTED